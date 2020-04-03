import openpyxl
#pip install --user -U openpyxl==2.6.2


wb = openpyxl.load_workbook('employee_data.xlsx')    
sheet = wb.active

def input_data():                                                   #This function will take input when a new entry is to be added
    first_name=input("Please enter the first name: ")
    last_name=input("Please enter the last name: ")
    city=input("Please enter the city name:")
    zip=input("Please enter a valid US zip code:")
    #Missing: Need to implement some type of validation here                  
    phone=input("Please enter a valid US phone number:")
    #Missing:Need to implement some type of validation here 
    add_data(first_name,last_name,city,int(zip),phone)

def add_data(first_name,last_name,city,zip,phone):              #This function will add the data in the spreadsheet
    emp_id=sheet.cell(row=sheet.max_row, column=1).value
    sheet.append((emp_id+1,first_name,last_name,city,zip,phone))
    wb.save('employee_data.xlsx')

def del_data(del_emp_id):                                       #Function to delete data
    del_row=0
    for i in range(2,sheet.max_row):
        if(sheet.cell(row=i,column=1).value==int(del_emp_id)):   #Searching if the entry is present in the data or not
            del_row=i
            break
    if del_row==0:
        print("No such entry found")
    else:
        sheet.delete_rows(del_row)
        print("Entry deleted!")
    wb.save('employee_data.xlsx')

def search_data(search_emp_id):                                 #Function to search for an entry
    flag=False
    for i in range(2,sheet.max_row):
        if(sheet.cell(row=i,column=1).value==int(search_emp_id)):                       
            print("Entry Found!! Here are the details:")
            print("Employee ID: " + search_emp_id)
            print("First Name: " + sheet.cell(row=i,column=2).value)
            print("Last Name: " + sheet.cell(row=i,column=3).value)
            print("City: " + sheet.cell(row=i,column=4).value)
            print("Zip: " + sheet.cell(row=i,column=5).value)
            print("Phone number: " + sheet.cell(row=i,column=6).value)
            flag=True
    if (flag==False):
        print("Sorry, No data found with that employee ID.")

while True:                                     #Infinite Loop until the user asks to exit
    choice=input("What do you want to do:\n1. Add an Entry\n2.Remove an Entry\n3.Search for an entry----")
    if choice=="1":
        input_data()
    if choice == "2":
        del_emp_id = input("Enter the id of the employee you want to delete:")
        del_data(del_emp_id)
    if choice=="3":
        search_emp_id = input("Enter the id of the employee you want to search for: ")
        search_data(search_emp_id)

    cont=input("If you want to continue please enter 1: ")
    if cont != "1":
        break