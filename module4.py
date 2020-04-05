from tkinter import Tk, Button, Label, Entry, CENTER
import openpyxl

#module4--Search Employee


wb = openpyxl.load_workbook('employee_data_2.xlsx')               #This connected with the excel file
sheet = wb.active                                               #Connects with the active sheet


def search_data(search_emp_id):                                 #Function to search for an entry

    for i in range(2,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value==int(search_emp_id)):

            #message = ("Employee ID: " + search_emp_id +"\nFirst Name: " + sheet.cell(row=i,column=2).value+ "\nLast Name: " + sheet.cell(row=i,column=3).value +
            #"\nCity: " + sheet.cell(row=i,column=4).value + "\nZip: " + str(sheet.cell(row=i,column=5).value) + "\nPhone number: " + sheet.cell(row=i,column=6).value ) 
            #I was using above statement to send the data before, but now using a list to send all the data
                        
            data = [sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=4).value,sheet.cell(row=i,column=5).value,sheet.cell(row=i,column=6).value]
            #Above list contains all the data of the employee if an entry is found in the database

            return data
            
    raise KeyError              #If the programs comes out of the for loop which will mean the entry is not found then it will raise an exception

def searchInstance():

    root1 = Tk()
    root1.title("Search Employee")
    root1.geometry('700x600')
    root1.configure(background ='#0084FF')

    lb2= Label(root1,text="", bg = '#0084FF' ,fg="white", font="Futura 12 bold")            #This label will display confirmation message
    lb2.place(x=280,y=420)
    
    Label (root1, text ="Search Entry", bg = '#0084FF' ,fg="white", font="Futura 20 bold").place(x =350, y = 100, anchor = CENTER)      #Heading
    Label(root1,text="Enter employee ID:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=200,y=200,anchor=CENTER)           #Label for text 
    user_entry = Entry (root1, width=20, bg="white")                                                                                    #Text box
    user_entry.place(x =350, y = 200, anchor = CENTER)

    def clicked():                                                              #When the button is pressed
        emp_id = user_entry.get()                                               #This will get employee id from the text box

        try:                                                                            #try-except for catching a exception if the emp id is not found in the database
            data = search_data(emp_id)
            message = f"Employee ID: {emp_id}\nFirst Name: {data[0]}\nLast Name: {data[1]}\nCity: {data[2]}\nZip: {data[3]}\nPhone number: {data[4]}" #adding the data to a message if the entry is found
        except KeyError:            #This exception will occur if no entry is found in the database 
            message="No data found with that employee id"         
            
        lb2.configure(text=message)
    Button (root1, text="SEARCH", width=60, command= clicked).place(x =350, y = 300, anchor = CENTER)

    return root1.mainloop()

#To add an event to the button follow the format below and add a function after the width parameter
#Button (root1, text="SEARCH", width=60, command = {Function}).place(x =350, y = 300, anchor = CENTER)
