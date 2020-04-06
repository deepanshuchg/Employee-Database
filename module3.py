from tkinter import Tk, Button, Label, Entry, CENTER, END
import openpyxl, re
#pip install --user -U openpyxl==2.6.2


wb = openpyxl.load_workbook('employee_data_2.xlsx')               #This connects with the excel file
sheet = wb.active                                               #Connects with the active sheet
#module3--Adding new entry   


def isValid(s):                                                 #Validating Phone number
	Pattern = re.compile(r'\d{3}-\d{3}-\d{4}')             #Regex for US style phone number
	return Pattern.match(s) 
    

def add_data(first_name,last_name,state,zip,phone):                     #This function will add the data in the spreadsheet
    emp_id=sheet.cell(row=sheet.max_row, column=1).value                #Getting the last employee id
    sheet.append((emp_id+1,first_name,last_name,state,zip,phone))
    wb.save('employee_data_2.xlsx')                                        #Saving the file


def newInstance():

    root1 = Tk()
    root1.title("New Employee")
    root1.geometry('700x600')
    root1.configure(background ='#0084FF')

    lb1= Label (root1, text ="New Entry", bg = '#0084FF' ,fg="white", font="Futura 20 bold")            #Heading
    lb1.place(x =350, y = 100, anchor = CENTER)

    lb2= Label(root1,text="", bg = '#0084FF' ,fg="white", font="Futura 20 bold")            #This label will display message when the data is added succesfully
    lb2.place(x=200,y=420)

    Label(root1,text="First Name:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=230,y=200,anchor=CENTER)      #Label for text "First Name"
    fname_entry = Entry(root1, width=20, bg="white")                                                                        #Text box for First Name
    fname_entry.place(x =360, y = 200, anchor = CENTER)

    Label(root1,text="Last Name:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=230,y=230,anchor=CENTER)
    lname_entry = Entry(root1, width=20, bg="white")
    lname_entry.place(x =360, y = 230, anchor = CENTER)

    Label(root1,text="State:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=230,y=260,anchor=CENTER)
    state_entry = Entry(root1, width=20, bg="white")
    state_entry.place(x =360, y = 260, anchor = CENTER)

    Label(root1,text="Zip Code:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=230,y=290,anchor=CENTER)
    zip_entry = Entry(root1, width=20, bg="white")
    zip_entry.place(x =360, y = 290, anchor = CENTER)

    Label(root1,text="Phone Number:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=230,y=320,anchor=CENTER)
    phone_entry = Entry(root1, width=20, bg="white")
    phone_entry.insert(0,"999-999-9999")                            #Putting the format in which phone number is to be added. Tkinter doesn't support placeholder
    phone_entry.place(x =360, y = 320, anchor = CENTER)

    def clicked():                                               #Function when the button is pressed

        first_name = fname_entry.get()                           #This section takes the input values from the text boxes
        last_name = lname_entry.get()
        state = state_entry.get()
        zip = zip_entry.get()
        phone = phone_entry.get()

        zip_entry.delete(0, END)
        phone_entry.delete(0, END)

        if len(zip)==5 and isValid(phone):                             #This will validate zip code and phone number are valid
            fname_entry.delete(0, END)                               #Clear the values of the text boxes   
            lname_entry.delete(0, END)
            state_entry.delete(0, END)
            
            add_data(first_name,last_name,state,int(zip),phone)        #Call the function to add values
            lb2.configure(text= "Entry added!")                   #Display confirmation message
        
        else:
            lb2.configure(text="Zip code or Phone number is invalid. Please enter the data again.", font="Futura 12 bold")


    Button (root1, text="Press to add a new Entry", width=60, command = clicked).place(x =350, y = 390, anchor = CENTER)            #Button object

    return root1.mainloop()