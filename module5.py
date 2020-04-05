from tkinter import Tk, Button, Label, Entry, CENTER
import openpyxl

#module5-- Deleting an entry

wb = openpyxl.load_workbook('employee_data_2.xlsx')               #This connected with the excel file
sheet = wb.active                                               #Connects with the active sheet

def del_data(del_emp_id):                                       #Function to delete data
    del_row=0                                                   #Variable to store the row to be deleted
    for i in range(2,sheet.max_row+1):
        if(sheet.cell(row=i,column=1).value==int(del_emp_id)):   #Searching if the entry is present in the data or not
            del_row = i
            break

    if del_row==0:                                                 #This section can be done with try/except like done in module4
        return("No such entry found")                              #But for now I haven't updated it
    else:
        sheet.delete_rows(del_row)
        wb.save('employee_data_2.xlsx')
        return("Entry deleted!")
    

def deleteInstance():

    root1 = Tk()
    root1.title("Delete Employee")
    root1.geometry('700x600')
    root1.configure(background ='#0084FF')
    
    Label (root1, text ="Delete Entry", bg = '#0084FF' ,fg="white", font="Futura 20 bold").place(x =350, y = 100, anchor = CENTER)      #Heading

    lb2= Label(root1,text="", bg = '#0084FF' ,fg="white", font="Futura 20 bold")            #This label will display confirmation message
    lb2.place(x=280,y=420)

    Label(root1,text="Enter employee ID:",bg = '#0084FF' ,fg="white", font="Futura 12 bold").place(x=200,y=200,anchor=CENTER)       #Text "Enter Employee ID"
    user_entry = Entry (root1, width=20, bg="white")                                                                                #Text box
    user_entry.place(x =350, y = 200, anchor = CENTER)                                                                      

    def clicked():                                  #Function when button is clicked

        emp_id = user_entry.get()                     #This will get the data from the text box
        message = del_data(emp_id)                    #Calling del function which will return a message based on entry is found or not
        lb2.configure(text=message)

    Button (root1, text="DELETE", width=60,command=clicked).place(x =350, y = 300, anchor = CENTER)
    return root1.mainloop()